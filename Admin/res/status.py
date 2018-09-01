# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'status.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
import pipeline as pipe
import pandas as pd
import time
import matplotlib as mpl
import matplotlib.pyplot as plt
import re
import pyperclip
import openpyxl
from openpyxl.styles import Font as xlFont
from openpyxl.styles import Alignment as xlAlign
from openpyxl.styles import Border as xlBoarder
from openpyxl.styles import Side as xlSide
from openpyxl.styles import PatternFill as xlPattern
from openpyxl.styles import Color as xlColor
from openpyxl.chart import (
    PieChart,
    Reference
)
from openpyxl.chart.label import DataLabelList
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas


index = {
    'connected': 0,
    'sleep': 1,
    'Time': 2,
    'temper': 3
}
MAX_TIME = 5.0
boundTemp = {
    'gt' : 36.5,
    'lt' : 0.0
}
item = ['_id', 'name', 'temper', 'connected', 'sleep', 'WiFi', 'Time']
field = {
    '_id': 0,
    'connected': 3,
    'temper': 2,
    'name': 1,
    'sleep': 4,
    'WiFi': 5,
    'Time': 6
}
def isFloat(element):
    if type(element) == str and re.match("^(-)?\d+?\.\d+?$", element) is None:
        return False
    else:
        return True

class Graph(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.setLayout(self.layout)

    def initUI(self):
        mpl.rcParams['font.size'] = 8.0
        mpl.rcParams['font.family'] = 'NanumGothic'
        self.fig = []
        self.canvas = []
        btnLayout = []
        for i in range(4):
            self.fig.append(plt.Figure())
            self.canvas.append(FigureCanvas(self.fig[i]))
            if i % 2  == 0:
                btnLayout.append(QVBoxLayout())
            btnLayout[i//2].addWidget(self.canvas[i])

        self.layout = QHBoxLayout()
        for i in range(2):
            self.layout.addLayout(btnLayout[i])
        self.value = 10

    def drawGraph(self, idx, data):
        title = ['턱끈 연결 현황', '졸음 현황', '통신 연결 현황', '온도 상태']
        self.fig[idx].clear()
        ax = self.fig[idx].subplots()

        # 'Connect' Pie chart
        sizes, labels = self.dataProcessing(idx, data)

        ax.pie(sizes, autopct='%1.1f%%')
        ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
        ax.legend(title=title[idx], labels=labels)
        self.canvas[idx].draw()

    @staticmethod
    def dataProcessing(idx, data):
        sizes, labels = None, None
        if idx == index['connected']:
            sizes = data
            labels = ['connect', 'disconnect']
        elif idx == index['sleep']:
            sizes = data
            labels = ['sleep', 'not sleep']
        elif idx == index['Time']:
            sizes = [0, 0]
            for T in data.values:
                if time.time() - float(T) < MAX_TIME:
                    sizes[0] += 1 # connect
                else:
                    sizes[1] += 1 # disconnect
            labels = ['connect', 'disconnect']
        elif idx == index['temper']:
            sizes = [0, 0]
            for T in data.values:
                if boundTemp['lt'] < float(T) <= boundTemp['gt']:
                    sizes[0] += 1 # NORMAL
                else:
                    sizes[1] += 1 # DANGER
            labels = ['normal', 'danger']

        return sizes, labels

class Ui_Form(object):
    def setupUi(self, Form):
        pipe.db.changeCollection(pipe.info['dataDB'])
        Form.setObjectName("Form")
        Form.resize(921, 610)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(Form.sizePolicy().hasHeightForWidth())
        Form.setSizePolicy(sizePolicy)
        self.groupBox = QtWidgets.QGroupBox(Form)
        self.groupBox.setGeometry(QtCore.QRect(20, 20, 891, 581))
        font = QtGui.QFont()
        font.setFamily("나눔스퀘어 Bold")
        self.groupBox.setFont(font)
        self.groupBox.setObjectName("groupBox")
        self.workerTab = QtWidgets.QTabWidget(self.groupBox)
        self.workerTab.setGeometry(QtCore.QRect(210, 40, 661, 531))
        font = QtGui.QFont()
        font.setFamily("나눔스퀘어 Bold")
        self.workerTab.setFont(font)
        self.workerTab.setAutoFillBackground(True)
        self.workerTab.setTabShape(QtWidgets.QTabWidget.Triangular)
        self.workerTab.setTabBarAutoHide(True)
        self.workerTab.setObjectName("workerTab")
        self.List = QtWidgets.QWidget()
        font = QtGui.QFont()
        font.setKerning(True)
        self.List.setFont(font)
        self.List.setObjectName("List")
        self.gridLayout_4 = QtWidgets.QGridLayout(self.List)
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.gridLayout = QtWidgets.QGridLayout()
        self.gridLayout.setObjectName("gridLayout")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setSizeConstraint(QtWidgets.QLayout.SetDefaultConstraint)
        self.verticalLayout.setObjectName("verticalLayout")
        self.tableTitle = QtWidgets.QLabel(self.List)
        font = QtGui.QFont()
        font.setFamily("나눔스퀘어 ExtraBold")
        font.setPointSize(19)
        font.setBold(True)
        font.setWeight(75)
        self.tableTitle.setFont(font)
        self.tableTitle.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.tableTitle.setAlignment(QtCore.Qt.AlignCenter)
        self.tableTitle.setObjectName("tableTitle")
        self.verticalLayout.addWidget(self.tableTitle)
        self.gridLayout.addLayout(self.verticalLayout, 0, 0, 1, 1)
        self.tableWidget = QtWidgets.QTableWidget(self.List)
        font = QtGui.QFont()
        font.setFamily("나눔바른고딕")
        self.tableWidget.setFont(font)
        self.tableWidget.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.tableWidget.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(7)
        self.tableWidget.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(6, item)
        self.gridLayout.addWidget(self.tableWidget, 1, 0, 1, 1)
        self.gridLayout_4.addLayout(self.gridLayout, 0, 0, 1, 1)
        self.workerTab.addTab(self.List, "")
        '''
        self.Status = QtWidgets.QWidget()
        self.Status.setObjectName("Status")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.Status)
        self.gridLayout_3.setObjectName("gridLayout_3")
        '''
        self.graph = Graph()
        font = QtGui.QFont()
        font.setFamily("나눔스퀘어 Bold")
        self.graph.setFont(font)
        self.workerTab.addTab(self.graph, "")
        self.searchTitle = QtWidgets.QLabel(self.groupBox)
        self.searchTitle.setGeometry(QtCore.QRect(20, 110, 121, 31))
        font = QtGui.QFont()
        font.setFamily("나눔고딕 ExtraBold")
        font.setPointSize(12)
        font.setBold(False)
        font.setWeight(50)
        self.searchTitle.setFont(font)
        self.searchTitle.setObjectName("searchTitle")
        self.lineEdit = QtWidgets.QLineEdit(self.groupBox)
        self.lineEdit.setGeometry(QtCore.QRect(20, 180, 141, 21))
        self.lineEdit.setObjectName("lineEdit")
        self.searchButton = QtWidgets.QPushButton(self.groupBox)
        self.searchButton.setGeometry(QtCore.QRect(20, 210, 61, 21))
        self.printAllButton = QtWidgets.QPushButton(self.groupBox)
        self.printAllButton.setGeometry(QtCore.QRect(90, 210, 61, 21))
        font = QtGui.QFont()
        font.setFamily("나눔스퀘어 Bold")
        self.searchButton.setFont(font)
        self.searchButton.setObjectName("searchButton")
        self.printAllButton.setFont(font)
        self.printAllButton.setObjectName("printAllButton")
        self.exportButton = QtWidgets.QPushButton(self.groupBox)
        self.exportButton.setGeometry(QtCore.QRect(20, 280, 61, 21))
        font = QtGui.QFont()
        font.setFamily("나눔스퀘어 Bold")
        self.exportButton.setFont(font)
        self.exportButton.setObjectName("exportButton")
        self.copyButton = QtWidgets.QPushButton(self.groupBox)
        self.copyButton.setGeometry(QtCore.QRect(90, 280, 61, 21))
        font = QtGui.QFont()
        font.setFamily("나눔스퀘어 Bold")
        self.copyButton.setFont(font)
        self.copyButton.setObjectName("copyButton")
        self.excelTitle = QtWidgets.QLabel(self.groupBox)
        self.excelTitle.setGeometry(QtCore.QRect(20, 240, 131, 31))
        font = QtGui.QFont()
        font.setFamily("나눔고딕 ExtraBold")
        font.setPointSize(12)
        font.setBold(False)
        font.setWeight(50)
        self.excelTitle.setFont(font)
        self.excelTitle.setObjectName("excelTitle")
        self.ListBox = QtWidgets.QComboBox(self.groupBox)
        self.ListBox.setGeometry(QtCore.QRect(60, 150, 101, 21))
        font = QtGui.QFont()
        font.setFamily("나눔스퀘어 Bold")
        self.ListBox.setFont(font)
        self.ListBox.setObjectName("ListBox")
        self.ListBox.addItem("")
        self.ListBox.addItem("")
        self.ListBox.addItem("")
        self.ListBox.addItem("")
        self.ListBox.addItem("")
        self.ListBox.addItem("")
        self.ListBox.addItem("")
        self.method = QtWidgets.QLabel(self.groupBox)
        self.method.setGeometry(QtCore.QRect(20, 140, 41, 41))
        font = QtGui.QFont()
        font.setFamily("나눔고딕")
        font.setBold(True)
        font.setWeight(75)
        self.method.setFont(font)
        self.method.setObjectName("method")
        self.Logo = QtWidgets.QLabel(self.groupBox)
        self.Logo.setGeometry(QtCore.QRect(20, 20, 161, 91))
        self.Logo.setText("")
        self.Logo.setTextFormat(QtCore.Qt.RichText)
        self.Logo.setPixmap(QtGui.QPixmap(":/logoIMG/logo.png"))
        self.Logo.setScaledContents(True)
        self.Logo.setObjectName("Logo")

        self.retranslateUi(Form)
        self.workerTab.setCurrentIndex(1)
        QtCore.QMetaObject.connectSlotsByName(Form)
        Form.setTabOrder(self.ListBox, self.lineEdit)
        Form.setTabOrder(self.lineEdit, self.searchButton)
        Form.setTabOrder(self.searchButton, self.printAllButton)
        Form.setTabOrder(self.printAllButton, self.exportButton)
        Form.setTabOrder(self.exportButton, self.copyButton)
        Form.setTabOrder(self.copyButton, self.workerTab)
        Form.setTabOrder(self.workerTab, self.tableWidget)

        self.getDataFromDatabase()

        self.databaseTimer = QtCore.QTimer(Form)
        self.databaseTimer.timeout.connect(self.getDataFromDatabase)
        self.databaseTimer.start(1500)

        self.searchButton.clicked.connect(self.search)
        self.printAllButton.clicked.connect(self.printAll)
        self.copyButton.clicked.connect(self.copyToClip)
        self.exportButton.clicked.connect(self.makeExcel)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "작업자 통합 관리 시스템"))
        self.groupBox.setTitle(_translate("Form", "Status"))
        self.tableTitle.setText(_translate("Form", "전체 작업자 상태"))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("Form", "직원번호"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("Form", "이름"))
        item = self.tableWidget.horizontalHeaderItem(2)
        item.setText(_translate("Form", "온도"))
        item = self.tableWidget.horizontalHeaderItem(3)
        item.setText(_translate("Form", "연결 상태"))
        item = self.tableWidget.horizontalHeaderItem(4)
        item.setText(_translate("Form", "졸음 상태"))
        item = self.tableWidget.horizontalHeaderItem(5)
        item.setText(_translate("Form", "연결 위치"))
        item = self.tableWidget.horizontalHeaderItem(6)
        item.setText(_translate("Form", "마지막 접속 시간"))
        self.workerTab.setTabText(self.workerTab.indexOf(self.List), _translate("Form", "Workers List"))
        self.workerTab.setTabText(self.workerTab.indexOf(self.graph), _translate("Form", "Workers Status"))
        self.searchTitle.setText(_translate("Form", "작업자 검색"))
        self.searchButton.setText(_translate("Form", "검색"))
        self.printAllButton.setText(_translate("Form", "전체 출력"))
        self.exportButton.setText(_translate("Form", "저장 하기"))
        self.copyButton.setText(_translate("Form", "복사 하기"))
        self.excelTitle.setText(_translate("Form", "엑셀 출력"))
        self.ListBox.setItemText(0, _translate("Form", "직원번호"))
        self.ListBox.setItemText(1, _translate("Form", "이름"))
        self.ListBox.setItemText(2, _translate("Form", "온도"))
        self.ListBox.setItemText(3, _translate("Form", "연결 상태"))
        self.ListBox.setItemText(4, _translate("Form", "졸음 상태"))
        self.ListBox.setItemText(5, _translate("Form", "연결 위치"))
        self.ListBox.setItemText(6, _translate("Form", "접속 시간"))
        self.method.setText(_translate("Form", "방  법"))

        self.findMethod = '_id'
        self.findQuery = ''


    def getDataFromDatabase(self):
        data = pipe.db.collection.find()
        dict = {}
        for key in data[0].keys():
            dict[key] = []
        for row in data:
            for key, val in row.items():
                dict[key].append(val)
        self.df = pd.DataFrame(data=dict)
        base = None
        if self.findQuery != '':
            query = self.find(self.findMethod, self.findQuery)
            pos = self.df[self.findMethod].isin(query)
            base = self.df.loc[pos]
        else:
            base = self.df
        self.drawTable(base)
        self.drawGraph()

    def find(self, method, query):
        result = []
        findData = query.split()
        for data in findData:
            for val in self.df[method].values:
                temp = str(val)
                if method == 'Time':
                    temp = str(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(val)))
                if method == 'connected':
                    if temp == '1': temp = 'True'
                    else: temp = 'False'
                if re.match(data, temp) is not None:
                    result.append(val)
        return result

    def drawGraph(self):
        self.graph.drawGraph(index['connected'], self.df['connected'].value_counts())
        self.graph.drawGraph(index['sleep'], self.df['sleep'].value_counts())
        self.graph.drawGraph(index['Time'], self.df['Time'])
        self.graph.drawGraph(index['temper'], self.df['temper'])

    def drawTable(self, df):
        self.tableWidget.setRowCount(len(df.index))
        self.tableWidget.setColumnCount(len(df.columns))

        for key, content in df.items():
            col = field[key]
            for row, val in enumerate(content):
                raw = val # time을 재변환 하는 것 보단 변수를 하나 더 사용하는 것이 이득이라 판단.
                if key == 'Time':
                    val = str(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(val)))
                if key == 'connected':
                    if val == '1': val = True
                    else: val = False
                self.tableWidget.setItem(row, col, QtWidgets.QTableWidgetItem(str(val)))
                if self.statusCheck(row, col, str(raw)):
                    self.tableWidget.item(row, col).setBackground(QtGui.QColor(255, 0, 0))

        self.tableWidget.resizeColumnsToContents()
        self.tableWidget.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)

    def statusCheck(self, row, col, data = ""):
        result = False
        if col == field['_id']: pass
        elif col == field['connected']:
            if data == '0':
                result = True
        elif col == field['temper']:
            if isFloat(data) or data.isdecimal():
                if not (boundTemp['lt'] < float(data) <= boundTemp['gt']):
                    result = True
        elif col == field['name']: pass
        elif col == field['sleep']:
            if data == 'True':
                result = True
        elif col == field['WiFi']: pass
        elif col == field['Time']:
            if isFloat(data) or data.isdecimal():
                if time.time() - float(data) > 5.0:
                    result = True
        else: print("Invalid Value Occur")
        return result

    def search(self):
        self.findMethod = item[self.ListBox.currentIndex()]
        self.findQuery = self.lineEdit.text()
        self.lineEdit.setText("")
        self.getDataFromDatabase()

    def printAll(self):
        self.findQuery = ""
        self.getDataFromDatabase()

    def copyToClip(self):
        rowSize = self.tableWidget.rowCount()
        colSize = self.tableWidget.columnCount()

        result = "직원번호, 이름, 온도, 연결 상태, 졸음 상태, 연결 위치, 접속 시간\r\n"

        for row in range(rowSize):
            for col in range(colSize):
                result += str(self.tableWidget.item(row, col).text())
                if col != colSize - 1: result += ', '
            result += '\r\n'
        pyperclip.copy(result)
        QMessageBox.question(pipe.window, '완료', '클립보드로 복사되었습니다.', QMessageBox.Yes, QMessageBox.Yes)

    def graphGenerator(self, sheet, pos, find, color, start):
        size, label = Graph().dataProcessing(index[find], self.df[find].value_counts())
        total = size[0] + size[1]
        data = [
            [label[0], ((size[0]/total) * 100)],
            [label[1], ((size[1]/total) * 100)]
        ]
        sheet.cell(row = 1, column = pos).font = color
        sheet.cell(row = 1, column = pos).value = 'label'
        sheet.cell(row = 1, column = pos + 1).font = color
        sheet.cell(row = 1, column = pos + 1).value = 'value'
        for idx, row in enumerate(data):
            sheet.cell(row=idx+2, column=pos).font = color
            sheet.cell(row=idx+2, column=pos+1).font = color
            sheet.cell(row=idx+2, column=pos).value = str(row[0])
            sheet.cell(row=idx+2, column=pos+1).value = int(row[1])
        pie = PieChart()
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        labels = Reference(sheet, min_col=pos, min_row=2, max_row=3)
        data = Reference(sheet, min_col=pos+1, min_row=1, max_row=3)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        title = 'None'
        if find == 'connected': title = '턱끈 연결 상태'
        elif find == 'Time': title = '통신 연결 상태'
        elif find == 'sleep': title = '졸음 상태'
        elif find == 'temper': title = '현장 온도 상태'
        pie.title = title
        pie.width = 9.5
        pie.height = 6

        sheet.add_chart(pie, start)


    def makeExcel(self):
        now = time.localtime()
        Font = {
            'title': xlFont(size=16, bold=True),
            'caption': xlFont(size=8),
            'header': xlFont(size=12, bold=True),
            'contents': xlFont(size=12),
            'blank': xlFont(color='FFFFFFFF')
        }

        Border = xlBoarder(left=xlSide(style='thin'),
                           right=xlSide(style='thin'),
                           top=xlSide(style='thin'),
                           bottom=xlSide(style='thin'))

        wb = openpyxl.Workbook()
        wb.create_sheet(index=0, title='요약')
        wb.create_sheet(index=1, title='리스트')
        wb.remove(wb['Sheet'])

        ############## sheet 요약 ##############
        sheet = wb['요약']
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.print_options.horizontalCentered = True
        sheet.print_options.verticalCentered = True
        sheet.print_area = 'B2:L26'
        self.graphGenerator(sheet, 15, 'connected', Font['blank'], 'B5')
        self.graphGenerator(sheet, 17, 'sleep', Font['blank'], 'B16')
        self.graphGenerator(sheet, 19, 'Time', Font['blank'], 'H5')
        self.graphGenerator(sheet, 21, 'temper', Font['blank'], 'H16')

        sheet['B2'].font = Font['title']
        sheet.merge_cells('B2:L2')
        sheet['B2'] = '['+str(now.tm_year)+'년 '+str(now.tm_mon)+'월 '+str(now.tm_mday)+'일] 현장 상태 보고서'
        sheet['B2'].alignment = xlAlign(horizontal ='center')

        sheet['B4'].font = Font['caption']
        sheet['B4'] = '출력 시간 : ' + str(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
        sheet['B4'].alignment = xlAlign(horizontal ='left')

        sheet['J4'].font = Font['caption']
        sheet['J4'] = '감독관 : '
        sheet['J4'].alignment = xlAlign(horizontal='right')

        sheet['L4'].font = Font['caption']
        sheet['L4'] = '(인)'
        sheet['L4'].alignment = xlAlign(horizontal='right')
        ############## sheet 리스트 ##############
        sheet = wb['리스트']
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.print_options.horizontalCentered = True
        sheet.print_options.verticalCentered = True
        sheet.print_area = 'B2:H26'

        sheet['B2'].font = Font['title']
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 8
        sheet.column_dimensions['E'].width = 11
        sheet.column_dimensions['F'].width = 11
        sheet.column_dimensions['G'].width = 16
        sheet.column_dimensions['H'].width = 24
        sheet.merge_cells('B2:H2')
        sheet['B2'] = '['+str(now.tm_year)+'년 '+str(now.tm_mon)+'월 '+str(now.tm_mday)+'일] 현장 상태 보고서'
        sheet['B2'].alignment = xlAlign(horizontal ='center')

        sheet['B4'].font = Font['caption']
        sheet['B4'] = '출력 시간 : ' + str(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))

        sheet['G4'].font = Font['caption']
        sheet['G4'] = '감독관 : '
        sheet['G4'].alignment = xlAlign(horizontal='right')

        sheet['H4'].font = Font['caption']
        sheet['H4'] = '(인)'
        sheet['H4'].alignment = xlAlign(horizontal='right')

        listHeader = ['직원번호', '이름', '온도', '연결 상태', '졸음 상태', '연결 위치', '접속 시간']
        base = 2

        headerColor = xlPattern(patternType='solid', fgColor=xlColor('FFFF00'))
        for idx in range(len(listHeader)):
            col = base + idx
            sheet.cell(row=5, column=col).font = Font['header']
            sheet.cell(row=5, column=col).alignment = xlAlign(horizontal='center')
            sheet.cell(row=5, column=col).value = listHeader[idx]
            sheet.cell(row=5, column=col).border = Border
            sheet.cell(row=5, column=col).fill = headerColor

        rowSize = self.tableWidget.rowCount()
        colSize = self.tableWidget.columnCount()

        base = {
            'row': 6,
            'col': 2
        }
        for row in range(rowSize):
            for col in range(colSize):
                sheet.cell(row=base['row']+row, column=base['col']+col).font = Font['contents']
                sheet.cell(row=base['row']+row, column=base['col']+col).alignment = xlAlign(horizontal='center')
                sheet.cell(row=base['row']+row, column=base['col']+col).border = Border
                data = str(self.tableWidget.item(row, col).text())
                if isFloat(data): data = float(data)
                elif data.isdecimal(): data = int(data)
                sheet.cell(row=base['row']+row, column=base['col']+col).value = data


        wb.save('['+str(now.tm_year)+'년 '+str(now.tm_mon)+'월 '+str(now.tm_mday)+'일] '
                +'현장 상태 보고서 ('+str(now.tm_hour)+'시'+str(now.tm_min)+'분'+').xlsx')

        QMessageBox.question(pipe.window, '완료', ' 엑셀 출력이 완료되었습니다.', QMessageBox.Yes, QMessageBox.Yes)

from res import LOGO_rc

if __name__ == "__main__":
    Ui_Form().makeExcel()
